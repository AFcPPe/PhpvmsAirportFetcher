
import sqlite3

import openpyxl as openpyxl

rwyLengthLimit = 1500
rwyLengthLimit = rwyLengthLimit/0.3048

workbook = openpyxl.Workbook()

# 制表Airports 0
sheetAirports = workbook.create_sheet("Airports",0)

sheetAirports.cell(1,1).value='icao'
sheetAirports.cell(1,2).value='iata'
sheetAirports.cell(1,3).value='name'
sheetAirports.cell(1,4).value='location'
sheetAirports.cell(1,5).value='country'
sheetAirports.cell(1,6).value='timezone'
sheetAirports.cell(1,7).value='hub'
sheetAirports.cell(1,8).value='lat'
sheetAirports.cell(1,9).value='lon'
sheetAirports.cell(1,10).value='ground_handling_cost'
sheetAirports.cell(1,11).value='fuel_100ll_cost'
sheetAirports.cell(1,12).value='fuel_jeta_cost'
sheetAirports.cell(1,13).value='fuel_mogas_cost'
sheetAirports.cell(1,14).value='notes'


line = 2

conn = sqlite3.connect('nd.db3')
cur = conn.cursor()
cur.execute("SELECT ID,AirportID,Length FROM Runways")
rwy = cur.fetchall()
cur.execute("SELECT ID,Name,ICAO,Latitude,Longtitude FROM Airports")
aips = cur.fetchall()

editedAips = {}

for each in aips:
    editedAips[str(each[0])] = each

count = 0
crtAip = {}
for each in rwy:
    if each[2]>=rwyLengthLimit:
        if str(each[1]) in crtAip:
            continue
        if not editedAips[str(each[1])][2].isalpha():
            continue
        crtAip[str(each[1])] = editedAips[str(each[1])]


        count+=1

for each in crtAip:

    sheetAirports.cell(line, 1).value = crtAip[each][2]
    sheetAirports.cell(line, 3).value = crtAip[each][1]
    sheetAirports.cell(line, 6).value = 'GMT'
    sheetAirports.cell(line, 8).value = crtAip[each][3]
    sheetAirports.cell(line, 9).value = crtAip[each][4]
    line+=1
print(count)
workbook.save('res.csv')